from copy import copy
from collections import OrderedDict
import matplotlib.pyplot as plt
import numpy as np
import mne
#import hypyp
import requests
import os
import PyQt5
import sys
import pickle
from openpyxl import load_workbook
import autoreject
from mne_icalabel import label_components
import subprocess
import copy
import my_utils
from my_utils import extract_trials, create_sub_epochs, AR_local_custom, ICA_autocorrect, AR_global_custom, link_eeg_to_behavioral_trials, get_channels_to_reject, create_time_locked_epochs, define_event_dictionary, create_erp_epochs


path = r"C:\Users\nicoucke\OneDrive - UGent\Desktop\Hyperscanning 1"
raw_path = r"C:\Users\nicoucke\OneDrive - UGent\Desktop\Hyperscanning 1\raw data"
prep_path = os.path.join(path, "time locked preprocessed")
log_path = os.path.join(path, "logs")


# loop through all data files
pair = 1
for root, dirs, files in os.walk(raw_path):
    for name in files:
        if name.endswith((".bdf")):
            print("processing file " + name)

            # define paths
            file_path = os.path.join(raw_path, name)
            log_folder_path = os.path.join(log_path, name)
            if not os.path.isdir(log_folder_path):
                os.makedirs(log_folder_path)

            split_name = name.split("P")
            pair = int ((int(split_name[1]) + 1) / 2)

            prep_filename = "".join(['pair_', str(pair)])
            # if the file already exists then not make it again
            continue_bool = False
            for root, dirs, files in os.walk(prep_path):
                if prep_filename in files:
                    continue_bool = True
            if continue_bool:
                continue

            # read in data
            raw = mne.io.read_raw_bdf(file_path, preload = True)
            channels_2 = raw.info.ch_names[:64]
            raw_2 = raw.pick_channels(channels_2)
            del raw
            raw = mne.io.read_raw_bdf(file_path, preload = True)
            channels_1 = raw.info.ch_names[76:140]
            raw_1 = raw.pick_channels(channels_1)
            del raw

            # note: we switch it up since in our setup, BOX1 in the eeg data is player 2 in the behavioral data

            # get the montage that we will use
            biosemi64_montage = mne.channels.make_standard_montage('biosemi64')

            # change the channel names in our epochs so that they are the same as the montage
            channels = biosemi64_montage.ch_names
            mapping_1 = {}
            mapping_2 = {}
            for i in range(len(channels)):
                mapping_1[raw_1.info['chs'][i]['ch_name']] = str(channels[i])
                mapping_2[raw_2.info['chs'][i]['ch_name']] = str(channels[i])
            mne.rename_channels(raw_1.info, mapping_1)
            mne.rename_channels(raw_2.info, mapping_2)

            raw_1.pick_channels(channels)
            raw_2.pick_channels(channels)


            # set the montage to the epochs
            raw_1.set_montage(biosemi64_montage)
            raw_2.set_montage(biosemi64_montage)


            #notch and band pass filtering
            raw_1.notch_filter(freqs = [50, 100, 150, 200, 250], filter_length='auto', method = 'fir', trans_bandwidth = 4, verbose = False)
            raw_1.filter(1, 45, verbose = False)

            raw_2.notch_filter(freqs = [50, 100, 150, 200, 250], filter_length='auto', method = 'fir', trans_bandwidth = 4, verbose = False)
            raw_2.filter(1, 45, verbose = False)

            raw = mne.io.read_raw_bdf(file_path)
            raw_events = mne.find_events(raw, shortest_event = 0)
            trials = extract_trials(raw_events)
            sfreq = raw.info['sfreq']
            old_event_id = {'Synchronous/Egalitarian': 2, 'Synchronous/LeaderFollower': 3, 'Synchronous/FollowerLeader': 4, 'Individual/Egalitarian': 5, 'Complementary/Egalitarian': 6, 'Complementary/LeaderFollower': 7, 'Complementary/FollowerLeader': 8}
           
            event_id = define_event_dictionary()
            events = create_erp_epochs(trials, sfreq, raw_events, old_event_id, event_id)

            
            # delete entries from event ids that are not present in the events
            
            keys_to_delete = []
            for  key, value in event_id.items():
                if not (value in events[:,2]):
                    keys_to_delete.append(key)
            for key in keys_to_delete:
                del event_id[key]
            
            
            raws = [raw_1, raw_2]
            """
            spliced_raws = []
            for raw_i in raws:
                cropped_segments = []
                # Loop through each segment and crop the raw data
                for segment in segments:
                    start_index, end_index = segment
                    
                    # Get the corresponding start and end times using the sample indices
                    start_time = raw_i.times[int(start_index-2)]
                    end_time = raw_i.times[int(end_index+2)]

                    # Crop the raw data using the obtained times
                    cropped_segment = raw_i.copy().crop(tmin=start_time, tmax=end_time)
                    cropped_segments.append(cropped_segment)

                # Concatenate the cropped segments back together
                spliced_raw = mne.concatenate_raws(cropped_segments)
                spliced_raws.append(spliced_raw)
            """
            spliced_raws = raws
            # reject very bad channels and then perform ICA:
            mappings = [mapping_1, mapping_2]

            raws_ICA_applied = []
            # interpolate bad channels but 
            map = 0
            num_reject_channels = []
            num_reject_components = []
            for spliced_raw in spliced_raws:
                
                # detect very bad channels
                rejected_channels = get_channels_to_reject(spliced_raw, events)
                channel_names = list(mappings[map].values())
                rejected_channel_names = [channel_names[i] for i in rejected_channels.tolist()]
                channels_to_reject = rejected_channel_names
                num_reject_channels.append(len(channels_to_reject))

                # interpolate them
                spliced_raw.info['bads'] = channels_to_reject
                spliced_raw.interpolate_bads()
                
                # set eeg reference after interpolation
                spliced_raw.set_eeg_reference(ref_channels='average', verbose = False)
     
                
                # Create a copy of the raw data and remove the interpolated channels from the copy
                #raw_copy = spliced_raw.copy().drop_channels(channels_to_reject)
                raw_copy = spliced_raw
                # skipped the above step because bad channels are excluded from ica anyway

                # Perform ICA on the copy without the interpolated channels
                ica = mne.preprocessing.ICA(n_components=35, random_state=97, max_iter=800)
                ica.fit(raw_copy)

                # reconstruct the data without the non-brain components
                ica_with_labels_fitted = label_components(raw_copy, ica, method="iclabel")
                ica_with_labels_component_detected = ica_with_labels_fitted["labels"]          
                #print(ica_with_labels_component_detected)
                ica_with_labels_probabilities = ica_with_labels_fitted["y_pred_proba"] 
                #print(ica_with_labels_probabilities)        
                excluded_idx_components = [idx for idx, label in enumerate(ica_with_labels_component_detected) if (label not in ["brain"] and  (ica_with_labels_probabilities[idx] > 0.9))]
                #excluded_idx_components = [idx for idx, label in enumerate(ica_with_labels_component_detected) if ((label in ["eye blink", "muscle artifact"]) and  (ica_with_labels_probabilities[idx] > 0.9))]
                num_reject_components.append(len(excluded_idx_components))
                raw_ica_applied = ica.apply(raw_copy.copy(), exclude=excluded_idx_components)


                # Interpolate the removed channels on the raw data with ICA applied
                raw_ica_applied.interpolate_bads(reset_bads=False)
                raws_ICA_applied.append(raw_ica_applied)
                map+=1

            # now epoch the data and perform autoreject        
            epochs = []

            for raw in raws_ICA_applied:
                epoch = mne.Epochs(raw, events, event_id, event_repeated = 'drop', on_missing = 'ignore', tmin=-1, tmax=1, preload=True, baseline=(0, 0))
                epoch.resample(sfreq=512)
                epochs.append(epoch)

            n_interpolates = np.array([1, 2, 4, 6])
            consensus_percs = np.linspace(0, 0.2, 5)

            cleaned_epochs_AR, dic_AR, bad_epochs_AR  = AR_local_custom(epochs, n_interpolates, consensus_percs,
                                                    strategy="union",
                                                    threshold=50.0,
                                                    verbose=False
            )

            # save the data
            storepath = os.path.join(prep_path,"pair_" + str(pair))
            with open(storepath, "wb") as output_file:
                pickle.dump(cleaned_epochs_AR, output_file, protocol=pickle.HIGHEST_PROTOCOL)

            # Path to your Excel file
            file_path = r"C:\Users\nicoucke\OneDrive - UGent\Desktop\Hyperscanning 1\time_locked_preprocessing_log.xlsx"
            # write line for first participant:
            for i in range(2):
                participant = 2*pair-1+i
                # Read the existing Excel file into a DataFrame

                percentage_bads = dic_AR['dyad']
                # The row number where you want to insert the new data (e.g., 5th row)
                target_row_number = participant+1

                # Load the existing workbook
                wb = load_workbook(file_path)

                # Select the active worksheet or a specific sheet
                ws = wb.active

                # Insert a new row (if you want to make space without overwriting existing data)
                ws.insert_rows(target_row_number)

                # Split your comma-separated string into values
                values = [pair, participant, num_reject_channels[i], num_reject_components[i], percentage_bads]

                # Assign values to each cell in the target row
                for col_index, value in enumerate(values, start=1):  # Starting index is 1 since Excel columns start from 1
                    ws.cell(row=target_row_number, column=col_index, value=value)

                # Save the workbook
                wb.save(file_path)

# Code to run script2.py at the end of script1.py
#subprocess.run(["python", "power_per_participant.py"])