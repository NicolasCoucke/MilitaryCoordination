from copy import copy
from collections import OrderedDict
import matplotlib.pyplot as plt
import numpy as np
import mne
#import hypyp
import requests
import os
import PyQt5
import sys
import pickle
from openpyxl import load_workbook
import autoreject
from mne_icalabel import label_components
import subprocess
import copy
import my_utils
from my_utils import extract_trials, create_sub_epochs, AR_local_custom, ICA_autocorrect, AR_global_custom, link_eeg_to_behavioral_trials, link_eeg_to_behavioral_trials_via_excel, get_channels_to_reject
import pandas as pd

path = r"C:\Users\nicoucke\OneDrive - UGent\Desktop\Hyperscanning 1"
raw_path = r"C:\Users\nicoucke\OneDrive - UGent\Desktop\Hyperscanning 1\raw data"
prep_path = os.path.join(path, "preprocessed data")
log_path = os.path.join(path, "logs")

excel_path = r"C:\Users\nicoucke\OneDrive - UGent\Desktop\Hyperscanning 1\match_excel_BU_BU.xlsx"

excel_path_cleaned = r"C:\Users\nicoucke\OneDrive - UGent\Desktop\Hyperscanning 1\match_excel_BU_BU_cleaned.xlsx"



def update_excel_with_eeg_times_for_pair(excel_path_cleaned, trials, sfreq, pair):
    """
    Updates the existing Excel file with the begin and end times in EEG for each trial, 
    specifically for the provided pair.

    Parameters:
    - excel_path (str): Path to the existing Excel file.
    - trials (ndarray): Array containing trials information with begin and end samples.
    - sfreq (float): Sampling frequency of the EEG data.
    - pair (str or int): The specific pair identifier to update in the Excel file.
    """
    # Load the workbook using openpyxl to manipulate visibility
    workbook = load_workbook(excel_path_cleaned)
    sheet_name = 'CleanedData'

    # Ensure at least one sheet is visible
    print(workbook.sheetnames)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

    # Ensure the sheet is visible
    if not workbook[sheet_name].sheet_state == 'visible':
        workbook[sheet_name].sheet_state = 'visible'

     # Load the existing Excel file into a pandas DataFrame
    column_names = ['Pair',	'Condition', 'Trial', 'CompletionTime',	'EEGIndex',	'EEGTime',	'EEG_Begin_Time', 'EEG_End_Time']


    excel_data = pd.read_excel(excel_path_cleaned, sheet_name=sheet_name, names = column_names)
    print(excel_data)
    # Filter rows for the specific pair
    pair_data = excel_data[excel_data['Pair'] == pair].copy()

    # Loop through each row in the filtered data to find the relevant trials and update with EEG times
    for idx, row in pair_data.iterrows():
        condition = row['Condition']
        eeg_index = row['EEGIndex']

        # If EEG index is NaN, skip the row
        if pd.isna(eeg_index):
            continue

        # Convert EEG index to integer for indexing
        eeg_index = int(eeg_index)

        # Filter the trials for the specific condition
        event_id = {'Synchronous/Egalitarian': 2, 'Synchronous/LeaderFollower': 3, 
                    'Synchronous/FollowerLeader': 4, 'Individual': 5, 
                    'Complementary/Egalitarian': 6, 'Complementary/LeaderFollower': 7, 
                    'Complementary/FollowerLeader': 8}

        condition_value = event_id.get(condition)

        # Filter the trials based on success and condition
        successes = np.array(trials[:, 4], dtype=int)
        segments = trials[(successes == 1) & (trials[:, 3] == condition_value), :]

        # Ensure EEG index is within the bounds of available segments
        if eeg_index < len(segments):
            # Get the begin and end samples in EEG time
            begin_sample = segments[eeg_index, 0]
            end_sample = segments[eeg_index, 1]

            # Convert samples to time
            begin_time = begin_sample / sfreq
            end_time = end_sample / sfreq

            # Update the DataFrame with the new times
            pair_data.at[idx, 'EEG_Begin_Time'] = begin_time
            pair_data.at[idx, 'EEG_End_Time'] = end_time
        else:
            print(f"Warning: EEG index {eeg_index} is out of bounds for condition {condition}.")

    # Update the original Excel data with the modified data for this pair
    excel_data.update(pair_data)

    # Save the updated Excel file back
    with pd.ExcelWriter(excel_path_cleaned, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        excel_data.to_excel(writer, sheet_name=sheet_name, index=False)



import pandas as pd
from openpyxl import load_workbook

def load_clean_and_save_excel(file_path, output_path):
    """
    Load an Excel file, remove rows that contain column names or are completely empty,
    concatenate all sheets, and save the cleaned data back to a new Excel file.

    :param file_path: Path to the Excel file to be cleaned.
    :param output_path: Path to the output Excel file where the cleaned data will be saved.
    """
    # Load the entire Excel file
    xls = pd.ExcelFile(file_path)

    # List to hold DataFrames for each sheet
    cleaned_sheets = []

    for sheet_name in xls.sheet_names:
        print(f"Processing sheet: {sheet_name}")

        # Read the entire sheet into a DataFrame without setting a header initially
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # Detect the first row that is not the header
        header_row = df[df.notna().all(axis=1)].index[0]

        # Set the columns to the detected header row and drop the header row from the data
        df.columns = df.iloc[header_row]
        df = df.drop(header_row).reset_index(drop=True)

        # Drop any rows that are identical to the header (these are repeated headers)
        df = df[~df.eq(df.columns).all(axis=1)]

        # Remove rows that are completely empty
        df = df.dropna(how='all')

        # Add cleaned DataFrame to the list
        cleaned_sheets.append(df)

    # Concatenate all DataFrames from different sheets
    all_data = pd.concat(cleaned_sheets, ignore_index=True)

    # Save the concatenated DataFrame to the output Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        all_data.to_excel(writer, sheet_name='CleanedData', index=False)

    print(f"Cleaned data saved to: {output_path}")

# Usage example
load_clean_and_save_excel(excel_path, excel_path_cleaned)



# Usage example
# Usage example:
# loop through all data files
pair = 1
for root, dirs, files in os.walk(raw_path):
    for name in files:
        if name.endswith((".bdf")):
            print("processing file " + name)

            # define paths
            file_path = os.path.join(raw_path, name)
            log_folder_path = os.path.join(log_path, name)
            if not os.path.isdir(log_folder_path):
                os.makedirs(log_folder_path)

            split_name = name.split("P")
            pair = int ((int(split_name[1]) + 1) / 2)

            prep_filename = "".join(['pair_', str(pair)])
        

            # extract trials 
            raw = mne.io.read_raw_bdf(file_path)
            events = mne.find_events(raw, shortest_event = 0)
            trials = extract_trials(events)

            print('trialshape ' + str(np.shape(trials)))
            # only keep trials for which we found behavioral data
            #trials = trials[np.where(~np.isnan(trials[:,4]))[0],:]
            
            # link the eeg and behavioral data
            with open(r"C:\Users\nicoucke\OneDrive - UGent\Desktop\Hyperscanning 1\behavioral data\Behavioral_Dataframe.pickle", "rb") as input_file:
                data_dictionary = pickle.load(input_file)
            sfreq = raw.info['sfreq']

            # only get successful trials
            successes = np.array(trials[:,4], dtype = int)
            segments = trials[successes == 1,:]
            #segments = segments[:,:2]
            #new_trial = [trial_begin, trial_end, trial_counter, condition, success]
            updated_trials = segments
            print('updated_trialshape ' + str(np.shape(updated_trials)))



            update_excel_with_eeg_times_for_pair(excel_path_cleaned, trials, sfreq, pair)



